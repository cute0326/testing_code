import re
from datetime import datetime,timedelta
import random
from openpyxl import load_workbook, Workbook
import os
import requests
from xml.etree import ElementTree
from time import sleep

def test_exist_or_not(test_phone_number):

    value = ''
    province = ''
    city = ''

    url = 'http://www2.ip138.com/xml/mobile_yisou.asp?mobile=' + test_phone_number
    response = requests.get(url)
    # response = requests.get('http://www.ip138.com:8080/search.asp?action=mobile&mobile=18310983748')
    readable_content = response.content.decode('utf-8')

    print(readable_content)
    try:
        value = readable_content.split('"')[1]
        print(value)
    except:
        pass

    try:
        province_city = readable_content.split('zone>')[1][0:-2]
        # print(province_city)
        province = province_city.split(' ')[0]
        city = province_city.split(' ')[1]
        print(province)
        print(city)
    except:
        pass

    return(province, city)


def generate_number_from_random():

    generate_number = ''

    for i in range(11):
        random_number = random.choice([1,2,3,4,5,6,7,8,9,0])
        generate_number += str(random_number)

    return generate_number

def phone_number(generate_number):
    phone_pattern = re.compile('((13[0-9])|(14[5|7])|(15([0-3]|[5-9]))|(18[3,0,5-9]))')

    result = re.match(phone_pattern, generate_number)

    if(result != None):
        return generate_number
    else:
        return None


def generete_excel_file(excel_name = 'phone_number.xlsx'):

    excel_file = Workbook()
    sheet_page = excel_file.create_sheet('电话号码过滤结果', index=0)

    sheet_page.cell(row=1, column=1).value = '电话号码'
    sheet_page.cell(row=1, column=2).value = '省'
    sheet_page.cell(row=1, column=3).value = '市'

    excel_file.save('./' + excel_name)


def write_date_to_excel(row_num, column_num, data):

    sheet_page = excel_file.get_sheet_by_name('电话号码过滤结果')

    sheet_page.cell(row=row_num, column=column_num).value = data

    excel_file.save('./phone_number.xlsx')

if __name__ == '__main__':

    start_time = datetime.now()

    generete_excel_file()
    print('have generated the excel file')

    excel_file = load_workbook('./phone_number.xlsx')

    column_num = 1
    row_num = 2

    while True:

        a = generate_number_from_random()

        if(a != None):
            b = phone_number(a)

        if(b != None):
            print(b)
            write_date_to_excel(row_num, column_num, b)
            row_num += 1
        else:
            pass

        end_time = datetime.now()

        if(end_time - start_time > timedelta(minutes=1)):
            break

    excel_file.save('./phone_number.xlsx')
    total_rows = row_num
    sheet_page = excel_file.get_sheet_by_name('电话号码过滤结果')

    for row_num in range(2,total_rows):

        excel_file = load_workbook('./phone_number.xlsx')
        sheet_page = excel_file.get_sheet_by_name('电话号码过滤结果')

        test_phone_number = sheet_page.cell(row=row_num, column=1).value
        print('test phone number is ' + str(test_phone_number))

        province, city = test_exist_or_not(test_phone_number)
        print(province, city)

        sheet_page.cell(row=row_num, column=2).value = province
        sheet_page.cell(row=row_num, column=3).value = city

        excel_file.save('./phone_number.xlsx')

        sleep(60 * 1)